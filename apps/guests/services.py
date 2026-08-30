import csv
import io
import re

from openpyxl import load_workbook

from apps.guests.models import Guest


GUEST_CATEGORY_LOOKUP = {
    'vip': Guest.Category.VIP,
    'famille': Guest.Category.FAMILY,
    'family': Guest.Category.FAMILY,
    'amis': Guest.Category.FRIENDS,
    'friends': Guest.Category.FRIENDS,
    'collegues': Guest.Category.COLLEAGUES,
    'colleagues': Guest.Category.COLLEAGUES,
    'temoins': Guest.Category.WITNESSES,
    'witnesses': Guest.Category.WITNESSES,
    'parents': Guest.Category.PARENTS,
    'autres': Guest.Category.OTHER,
    'other': Guest.Category.OTHER,
}


def _clean_header(value):
    return str(value or '').strip()


def _normalize_phone(value):
    return re.sub(r'\s+', '', str(value or '').strip())


def _parse_csv(uploaded_file):
    decoded = uploaded_file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(decoded))
    headers = [header for header in reader.fieldnames or [] if header]
    rows = []
    for row in reader:
        rows.append({_clean_header(key): str(value or '').strip() for key, value in row.items() if key})
    return headers, rows


def _parse_excel(uploaded_file):
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return [], []
    headers = [_clean_header(cell) for cell in values[0] if cell is not None]
    rows = []
    for line in values[1:]:
        row = {}
        for index, header in enumerate(headers):
            row[header] = str(line[index] or '').strip() if index < len(line) else ''
        if any(row.values()):
            rows.append(row)
    return headers, rows


def parse_guest_import_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        return _parse_csv(uploaded_file)
    return _parse_excel(uploaded_file)


def _map_category(value):
    normalized = str(value or '').strip().lower()
    return GUEST_CATEGORY_LOOKUP.get(normalized, Guest.Category.OTHER)


def _map_companions(value):
    if value in (None, ''):
        return 0
    try:
        return max(int(float(value)), 0)
    except (TypeError, ValueError):
        return 0


def import_guest_rows(*, event, rows, mapping, import_mode):
    summary = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}

    for index, source_row in enumerate(rows, start=2):
        payload = {}
        for field_name, column in mapping.items():
            payload[field_name] = source_row.get(column, '').strip() if column else ''

        full_name = payload.get('full_name', '').strip()
        whatsapp_number = _normalize_phone(payload.get('whatsapp_number', ''))
        if not full_name or not whatsapp_number:
            summary['skipped'] += 1
            summary['errors'].append(f'Ligne {index}: nom complet ou numero WhatsApp manquant.')
            continue

        defaults = {
            'full_name': full_name,
            'email': payload.get('email', ''),
            'category': _map_category(payload.get('category', '')),
            'allowed_companions': _map_companions(payload.get('allowed_companions', '')),
            'reserved_table': payload.get('reserved_table', ''),
            'notes': payload.get('notes', ''),
        }

        if import_mode == 'upsert':
            _, created = Guest.objects.update_or_create(
                event=event,
                whatsapp_number=whatsapp_number,
                defaults=defaults,
            )
            summary['created' if created else 'updated'] += 1
            continue

        if Guest.objects.filter(event=event, whatsapp_number=whatsapp_number).exists():
            summary['skipped'] += 1
            continue

        Guest.objects.create(event=event, whatsapp_number=whatsapp_number, **defaults)
        summary['created'] += 1

    return summary